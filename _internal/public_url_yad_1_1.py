import os

# Uncomment the following lines if you need to install packages automatically
os.system("pip install requests")
os.system("pip install openpyxl")
os.system("pip install yadisk")
os.system("pip install tqdm")

import requests
import openpyxl
# from tqdm import tqdm
from tqdm import tqdm, tqdm_gui, trange

import yadisk

# Function to get the public URL from Yandex Disk
def get_public_url(disk_link, access_token):
    # Extract the file ID from the Yandex Disk link
    file_id = disk_link.split('/')[-1]

    # Yandex Disk API endpoint for getting file information
    api_url = f"https://cloud-api.yandex.net/v1/disk/resources?path={file_id}&fields=public_url"

    headers = {
        'Authorization': f'OAuth {access_token}'
    }

    response = requests.get(api_url, headers=headers)

    if response.status_code == 200:
        data = response.json()
        public_url = data.get('public_url')
        if public_url:
            return public_url
        else:
            return "The file does not have a public URL."
    else:
        return f"Error: {response.status_code} - {response.text}"


def main():
    access_token = input("Enter your Yandex Disk access token: ")

    # Open the Excel file
    exl = openpyxl.load_workbook(r"urls.xlsx")
    sheet = exl.active

    # Iterate through all rows in the first column
    for row in tqdm(range(2, sheet.max_row + 1), desc='Processing:'):
        disk_link = sheet.cell(row=row, column=1).value

        if disk_link:
            # Get the public URL
            public_url = get_public_url(disk_link, access_token)
            print(f"Public URL for row {row}: {public_url}")

            # Write the public URL to the second column of the same row
            sheet.cell(row=row, column=2, value=public_url)
        else:
            print(f"No Yandex Disk link found in row {row}.")

    # Save the changes to the Excel file
    exl.save(r"urls.xlsx")
    print("Public URLs have been saved.")


if __name__ == "__main__":
    main()

input("Формирование ссылок public_url завершено. Нажмите Enter для закрытия окна")

# softy_plug